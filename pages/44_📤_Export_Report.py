"""
Module 44: Export & Report — DataScience Flow v9.5
Part of HMG Academy Ecosystem — Subsidiary of HMG Concepts
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from subscription import init_subscription_state, TIERS, tier_badge_html, get_tier_info, local_save_subscription
from security import (
    init_secure_session, verify_subscription_integrity,
    authenticate_subscription, check_rate_limit, check_export_limit,
    validate_upload, sanitise_string, sanitise_email,
    log_action, check_session_timeout,
    watermark_dataframe, BRAND_WATERMARK, BRAND_FOOTER,
)

# ═══ Security & Session Verification ═══
init_secure_session()
if not st.session_state.get("sub_authenticated"):
    st.warning("⚠️ Subscription required. Start your free trial from the Home page.")
    st.stop()
if not verify_subscription_integrity():
    st.error("🔒 Session integrity check failed. Please re-authenticate.")
    st.stop()
if check_session_timeout():
    st.warning("⏰ Session expired. Please re-authenticate.")
    st.stop()
check_rate_limit()

if not verify_subscription_integrity():
    st.error("🔒 Session integrity check failed. Please re-authenticate.")
    st.stop()
if check_session_timeout():
    st.warning("⏰ Session expired. Please re-authenticate.")
    st.stop()
check_rate_limit()



st.set_page_config(page_title="Export & Report | DataScience Flow", page_icon="📤", layout="wide")
st.markdown('<div style="font-size:2.2rem;font-weight:800;background:linear-gradient(135deg,#6C63FF,#00D9A7);-webkit-background-clip:text;-webkit-text-fill-color:transparent;">📤 EXPORT & REPORT</div>', unsafe_allow_html=True)
st.markdown("""**🧠 What this module does:** Export your cleaned data in 5+ formats: CSV, Excel, Styled Excel (colour-coded), Parquet, JSON, HTML EDA Report (Sweetviz), and a text summary with Data Dictionary.

**🎯 Why you need it:** After all your analysis and cleaning, you need to get your data out in the right format for sharing, presentation, or further processing.

**📖 How to use it:** Select format → configure options → download. Also generates a comprehensive HTML EDA report.""")

if st.session_state.get("df") is None:
    st.warning("⚠️ No dataset loaded. Please upload a file from the Home page sidebar.")
    st.stop()

dfc = st.session_state["df_cleaned"]
st.session_state.get("steps_done", set()).add("export")

st.markdown("### 📤 Export Cleaned Data")

fmt = st.selectbox("Export format", ["CSV", "Excel (.xlsx)", "Styled Excel (colour-coded)", "Parquet", "JSON", "HTML EDA Report (Sweetviz)"])

if fmt == "CSV":
    data = dfc.to_csv(index=False).encode()
    st.download_button("⬇️ Download CSV", data, "cleaned_data.csv", "text/csv")

elif fmt == "Excel (.xlsx)":
    import io
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        dfc.to_excel(w, index=False, sheet_name="Cleaned Data")
    st.download_button("⬇️ Download Excel", buf.getvalue(), "cleaned_data.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif fmt == "Styled Excel (colour-coded)":
    import io
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        dfc.to_excel(w, index=False, sheet_name="Cleaned Data")
        ws = w.sheets["Cleaned Data"]
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill(start_color="00d9a7", end_color="00d9a7", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = Font(bold=True, color="0d1117")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for i, cell in enumerate(row):
                if i % 2 == 0:
                    cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
    st.download_button("⬇️ Download Styled Excel", buf.getvalue(), "cleaned_data_styled.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif fmt == "Parquet":
    import io
    buf = io.BytesIO()
    dfc.to_parquet(buf, index=False)
    st.download_button("⬇️ Download Parquet", buf.getvalue(), "cleaned_data.parquet", "application/octet-stream")

elif fmt == "JSON":
    data = dfc.to_json(orient="records").encode()
    st.download_button("⬇️ Download JSON", data, "cleaned_data.json", "application/json")

elif fmt == "HTML EDA Report (Sweetviz)":
    st.info("Generating Sweetviz HTML report... (may take a moment)")
    try:
        import sweetviz as sv
        report = sv.analyze(dfc)
        import io, tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
            report.show_html(f.name, open_browser=False)
            with open(f.name, 'r') as rf:
                html_content = rf.read()
            os.unlink(f.name)
        st.download_button("⬇️ Download HTML EDA Report", html_content.encode(), "eda_report.html", "text/html")
    except Exception as e:
        st.error(f"Sweetviz error: {e}")

st.markdown("### 📝 Session Summary")
st.markdown(f"- **Dataset:** {st.session_state.get('filename', 'Unknown')}")
st.markdown(f"- **Rows:** {dfc.shape[0]:,} × **Columns:** {dfc.shape[1]}")
st.markdown(f"- **Audit events:** {len(st.session_state.get('ent_audit_log', []))}")
st.markdown(f"- **Steps completed:** {len(st.session_state.get('steps_done', set()))}")

# ═══════════════════════════════════════════════════════════════
# BRAND FOOTER — DataScience Flow v9.5 | HMG Academy
# ═══════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown("""
<div style="text-align:center; padding:16px 0; font-size:0.82rem; color:#A0A8B8;">
<b>DataScience Flow</b> v9.5 — Part of <b>HMG Academy</b> Ecosystem — Subsidiary of <b>HMG Concepts</b> (est. 2015)<br>
Built by <b>Adewale Samson Adeagbo</b> | 🔒 Secured Platform | 🏗️ HMG Academy · HMG Technologies · HMG Media · HMG Gospel<br>
<a href="https://hmgacademy.pages.dev" style="color:#6C63FF;">hmgacademy.pages.dev</a> ·
<a href="https://hmgconcepts.pages.dev" style="color:#6C63FF;">hmgconcepts.pages.dev</a> ·
<a href="https://cssadewale.pages.dev" style="color:#6C63FF;">cssadewale.pages.dev</a>
</div>
""", unsafe_allow_html=True)
